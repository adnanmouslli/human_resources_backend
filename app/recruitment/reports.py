# app/recruitment/reports.py
# تصدير تقارير التوظيف - Excel و PDF
# تقرير احترافي وكامل: يعرض كل بيانات الطلب بما فيها الإجابات الديناميكية،
# الخبرات السابقة (الأعمال السابقة)، وقرار التعيين النهائي.

import io
import json
from datetime import datetime, timedelta

from app.recruitment.models import (
    RecruitmentApplication,
    RecruitmentApplicationAnswer,
    RecruitmentFormSection,
    RecruitmentFormField,
    RecruitmentApplicationExperience,
    HiringDecision,
)


# ─────────────────────────────────────────────────────────────────
# مساعدات مشتركة
# ─────────────────────────────────────────────────────────────────

def _get_ordered_sections_with_fields():
    """
    إرجاع قائمة الأقسام النشطة مرتبة، وكل قسم مع حقوله النشطة مرتبة.
    [(section, [field, field, ...]), ...]
    """
    sections = (
        RecruitmentFormSection.query
        .filter_by(is_active=True)
        .order_by(RecruitmentFormSection.display_order)
        .all()
    )
    result = []
    for section in sections:
        fields = (
            RecruitmentFormField.query
            .filter_by(section_id=section.id, is_active=True)
            .order_by(RecruitmentFormField.display_order)
            .all()
        )
        result.append((section, fields))
    return result


def _get_ordered_fields():
    """
    إرجاع قائمة بكل الحقول النشطة مرتبة حسب القسم ثم display_order.
    تُستخدم لبناء أعمدة التقارير بشكل ديناميكي.
    """
    ordered_fields = []
    for _section, fields in _get_ordered_sections_with_fields():
        ordered_fields.extend(fields)
    return ordered_fields


def _get_applications(filters=None):
    """استرجاع الطلبات مع دعم الفلترة"""
    if filters is None:
        filters = {}

    query = RecruitmentApplication.query

    if filters.get('status'):
        query = query.filter(RecruitmentApplication.status == filters['status'])
    if filters.get('branch_id'):
        query = query.filter(RecruitmentApplication.branch_id == filters['branch_id'])
    if filters.get('department_id'):
        query = query.filter(RecruitmentApplication.department_id == filters['department_id'])
    if filters.get('date_from'):
        try:
            df = datetime.strptime(filters['date_from'], '%Y-%m-%d')
            query = query.filter(RecruitmentApplication.created_at >= df)
        except ValueError:
            pass
    if filters.get('date_to'):
        try:
            dt = datetime.strptime(filters['date_to'], '%Y-%m-%d')
            # نهاية اليوم لتضمين كل سجلات نفس التاريخ
            dt_end = dt + timedelta(days=1) - timedelta(seconds=1)
            query = query.filter(RecruitmentApplication.created_at <= dt_end)
        except ValueError:
            pass

    return query.order_by(RecruitmentApplication.created_at.desc()).all()


def _build_answers_map(application_ids):
    """
    بناء قاموس { application_id: { field_id: value } }
    باستعلام واحد لكل الطلبات (تجنب N+1).
    """
    if not application_ids:
        return {}

    answers = (
        RecruitmentApplicationAnswer.query
        .filter(RecruitmentApplicationAnswer.application_id.in_(application_ids))
        .all()
    )
    result = {}
    for ans in answers:
        result.setdefault(ans.application_id, {})[ans.field_id] = ans.value
    return result


def _build_experiences_map(application_ids):
    """
    بناء قاموس { application_id: [experiences] }
    باستعلام واحد.
    """
    if not application_ids:
        return {}

    exps = (
        RecruitmentApplicationExperience.query
        .filter(RecruitmentApplicationExperience.application_id.in_(application_ids))
        .order_by(RecruitmentApplicationExperience.application_id,
                  RecruitmentApplicationExperience.experience_order)
        .all()
    )
    result = {}
    for exp in exps:
        result.setdefault(exp.application_id, []).append(exp)
    return result


def _build_hiring_map(application_ids):
    """
    بناء قاموس { application_id: HiringDecision } باستعلام واحد.
    """
    if not application_ids:
        return {}

    decisions = (
        HiringDecision.query
        .filter(HiringDecision.application_id.in_(application_ids))
        .all()
    )
    return {d.application_id: d for d in decisions}


STATUS_LABELS = {
    'new': 'جديد',
    'under_review': 'قيد المراجعة',
    'interview': 'مقابلة',
    'accepted': 'مقبول',
    'rejected': 'مرفوض',
}

EMPLOYEE_TYPE_LABELS = {
    'permanent': 'دائم',
    'temporary': 'مؤقت',
}

# الأعمدة الفرعية لكل خبرة (عمل سابق)
EXP_SUB_HEADERS = [
    ('company_name', 'الشركة / العمل السابق'),
    ('company_field', 'مجال الشركة'),
    ('position', 'الوظيفة'),
    ('duration', 'مدة العمل'),
    ('hours_per_day', 'الساعات/اليوم'),
    ('salary', 'المرتب'),
    ('reason_for_leaving', 'سبب ترك العمل'),
]


def _fmt(value):
    """تنسيق قيمة للعرض النصي."""
    if value is None:
        return ''
    return str(value)


def _hiring_columns(decision):
    """قائمة [(label, value)] لبيانات قرار التعيين."""
    if not decision:
        return []
    rows = [
        ('المسمى الوظيفي (التعيين)', decision.job_title.title_name if decision.job_title else ''),
        ('فرع التعيين', decision.branch.name if decision.branch else ''),
        ('قسم التعيين', decision.department.name if decision.department else ''),
        ('المرتب', f"{decision.salary:.2f}" if decision.salary is not None else ''),
        ('مرتب فترة التجربة', f"{decision.probation_salary:.2f}" if decision.probation_salary is not None else ''),
        ('تاريخ المباشرة', decision.start_date.strftime('%Y-%m-%d') if decision.start_date else ''),
        ('نوع الموظف', EMPLOYEE_TYPE_LABELS.get(decision.employee_type, decision.employee_type or '')),
        ('نظام العمل', decision.work_system or ''),
        ('الوردية', decision.shift.name if decision.shift else ''),
        ('المهنة', decision.profession.name if decision.profession else ''),
        ('رقم البصمة', decision.fingerprint_id or ''),
        ('ساعات الدوام', decision.working_hours or ''),
        ('ملاحظات التعيين', decision.notes or ''),
        ('قرر بواسطة', decision.decider.username if decision.decider else ''),
        ('تاريخ القرار', decision.decided_at.strftime('%Y-%m-%d %H:%M') if decision.decided_at else ''),
    ]
    return rows


# ─────────────────────────────────────────────────────────────────
# تصدير Excel
# ─────────────────────────────────────────────────────────────────

def export_excel(filters=None):
    """
    تصدير بيانات الطلبات إلى Excel - تقرير كامل.
    الأعمدة:
      - معلومات أساسية (رقم، الحالة، الوظيفة المتقدم لها، الفرع، القسم،
        التقييم، سبب الرفض، المُدخِل، تاريخ التقديم)
      - الحقول الديناميكية النشطة (ما عدا experience_group)
      - الخبرات السابقة مُفرَّدة لكل خبرة
      - بيانات قرار التعيين الكاملة (للمقبولين)

    يعيد BytesIO object جاهز للإرسال كملف.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise RuntimeError("مكتبة openpyxl غير مثبتة")

    applications = _get_applications(filters)
    ordered_fields = _get_ordered_fields()

    app_ids = [a.id for a in applications]
    answers_map = _build_answers_map(app_ids)
    experiences_map = _build_experiences_map(app_ids)
    hiring_map = _build_hiring_map(app_ids)

    # تحديد الحد الأقصى لعدد الخبرات
    max_experiences = max(
        (len(exps) for exps in experiences_map.values()),
        default=0
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "طلبات التوظيف"
    ws.sheet_view.rightToLeft = True  # RTL

    # ───── أنماط ─────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    exp_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    hiring_fill = PatternFill(start_color="B7791F", end_color="B7791F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ───── بناء رأس الجدول ─────
    base_headers = [
        "#", "الحالة", "الوظيفة المتقدم لها", "الفرع", "القسم",
        "التقييم العام", "سبب الرفض", "أُدخل بواسطة", "تاريخ التقديم",
    ]
    base_count = len(base_headers)

    # أعمدة الحقول الديناميكية (ما عدا experience_group)
    dynamic_fields = [f for f in ordered_fields if f.field_type != 'experience_group']
    dynamic_headers = [f.label for f in dynamic_fields]

    # أعمدة الخبرات المفرَّدة
    exp_headers = []
    for i in range(1, max_experiences + 1):
        for _key, label in EXP_SUB_HEADERS:
            exp_headers.append(f"خبرة {i} - {label}")

    # أعمدة قرار التعيين
    hiring_labels = [lbl for lbl, _ in _hiring_columns(HiringDecision())]
    hiring_headers = [f"تعيين - {lbl}" for lbl in hiring_labels]

    headers = base_headers + dynamic_headers + exp_headers + hiring_headers

    # كتابة الرأس مع تلوين المجموعات
    dyn_start = base_count
    exp_start = dyn_start + len(dynamic_headers)
    hire_start = exp_start + len(exp_headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        if col_num - 1 >= hire_start:
            cell.fill = hiring_fill
        elif col_num - 1 >= exp_start:
            cell.fill = exp_fill
        else:
            cell.fill = header_fill

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # ───── بناء صفوف البيانات ─────
    for row_num, app in enumerate(applications, 2):
        ans = answers_map.get(app.id, {})
        exps = experiences_map.get(app.id, [])
        decision = hiring_map.get(app.id)

        row_data = [
            app.id,
            STATUS_LABELS.get(app.status, app.status),
            app.applied_position or '',
            app.branch.name if app.branch else '',
            app.department.name if app.department else '',
            app.evaluation_score if app.evaluation_score is not None else '',
            app.rejection_reason or '',
            app.creator.username if app.creator else '',
            app.created_at.strftime('%Y-%m-%d') if app.created_at else '',
        ]

        # قيم الحقول الديناميكية
        for field in dynamic_fields:
            row_data.append(_fmt(ans.get(field.id, '')))

        # قيم الخبرات المفرَّدة
        for i in range(max_experiences):
            if i < len(exps):
                exp = exps[i]
                for key, _label in EXP_SUB_HEADERS:
                    row_data.append(getattr(exp, key) or '')
            else:
                row_data.extend([''] * len(EXP_SUB_HEADERS))

        # قيم قرار التعيين
        if decision:
            for _label, value in _hiring_columns(decision):
                row_data.append(value)
        else:
            row_data.extend([''] * len(hiring_labels))

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_align
            cell.border = border

    # ضبط عرض الأعمدة
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────────────
# تصدير PDF - تخطيط تفصيلي لكل طلب (يضمن عرض كل البيانات)
# ─────────────────────────────────────────────────────────────────

def export_pdf(filters=None):
    """
    تصدير بيانات الطلبات إلى PDF مع دعم RTL.
    تخطيط تفصيلي: لكل طلب بطاقة كاملة تعرض:
      - المعلومات الأساسية
      - كل الإجابات مجمّعة حسب الأقسام
      - جدول الخبرات السابقة (الأعمال السابقة) كاملاً
      - بيانات قرار التعيين (للمقبولين)
    يعيد BytesIO object جاهز للإرسال كملف.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
            KeepTogether, HRFlowable,
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import arabic_reshaper
        from bidi.algorithm import get_display
    except ImportError as e:
        raise RuntimeError(f"مكتبة مطلوبة غير مثبتة: {e}")

    # ──── محاولة تسجيل خط عربي ────
    import os
    fonts_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts')
    arabic_font = 'Helvetica'  # fallback
    arabic_font_bold = 'Helvetica-Bold'
    for font_file in ['Amiri-Regular.ttf', 'Cairo-Regular.ttf', 'NotoSansArabic-Regular.ttf']:
        font_path = os.path.join(fonts_dir, font_file)
        if os.path.exists(font_path):
            font_name = font_file.replace('.ttf', '').replace('-Regular', '')
            pdfmetrics.registerFont(TTFont(font_name, font_path))
            arabic_font = font_name
            arabic_font_bold = font_name  # نفس الخط (لا يوجد bold منفصل غالباً)
            # محاولة تسجيل نسخة bold إن وُجدت
            bold_path = font_path.replace('-Regular', '-Bold')
            if os.path.exists(bold_path):
                bold_name = font_name + '-Bold'
                pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                arabic_font_bold = bold_name
            break

    def reshape(text):
        """تشكيل النص العربي للعرض الصحيح"""
        if text is None or text == '':
            return ''
        try:
            reshaped = arabic_reshaper.reshape(str(text))
            return get_display(reshaped)
        except Exception:
            return str(text)

    applications = _get_applications(filters)
    sections_with_fields = _get_ordered_sections_with_fields()

    app_ids = [a.id for a in applications]
    answers_map = _build_answers_map(app_ids)
    experiences_map = _build_experiences_map(app_ids)
    hiring_map = _build_hiring_map(app_ids)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=1.2 * cm, leftMargin=1.2 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title='تقرير طلبات التوظيف',
    )

    base_style = ParagraphStyle(
        'Arabic', fontName=arabic_font, fontSize=9, leading=13, alignment=2,  # RIGHT
    )
    label_style = ParagraphStyle(
        'ArabicLabel', fontName=arabic_font_bold, fontSize=9, leading=13,
        alignment=2, textColor=colors.HexColor('#1F4E79'),
    )
    value_style = ParagraphStyle(
        'ArabicValue', fontName=arabic_font, fontSize=9, leading=13, alignment=2,
    )
    main_title_style = ParagraphStyle(
        'ArabicMainTitle', fontName=arabic_font_bold, fontSize=18, leading=22,
        alignment=1, textColor=colors.HexColor('#1F4E79'),
    )
    app_title_style = ParagraphStyle(
        'ArabicAppTitle', fontName=arabic_font_bold, fontSize=12, leading=16,
        alignment=2, textColor=colors.white,
    )
    section_title_style = ParagraphStyle(
        'ArabicSection', fontName=arabic_font_bold, fontSize=10, leading=14,
        alignment=2, textColor=colors.HexColor('#1F4E79'),
    )
    th_style = ParagraphStyle(
        'ArabicTH', fontName=arabic_font_bold, fontSize=8, leading=11,
        alignment=1, textColor=colors.white,
    )
    td_style = ParagraphStyle(
        'ArabicTD', fontName=arabic_font, fontSize=8, leading=11, alignment=1,
    )

    def P(text, style=value_style):
        return Paragraph(reshape(text), style)

    story = []

    # ── ترويسة التقرير ──
    story.append(P("تقرير طلبات التوظيف", main_title_style))
    story.append(Spacer(1, 0.2 * cm))
    story.append(P(f"تاريخ الاستخراج: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   "
                   f"إجمالي الطلبات: {len(applications)}", base_style))
    story.append(Spacer(1, 0.4 * cm))

    if not applications:
        story.append(P("لا توجد طلبات مطابقة للفلاتر المحددة.", base_style))
        doc.build(story)
        output.seek(0)
        return output

    content_width = doc.width

    def kv_table(pairs):
        """جدول مفتاح-قيمة بعمودين (قيمة | تسمية) بمحاذاة RTL."""
        rows = []
        for label, value in pairs:
            rows.append([P(value, value_style), P(label, label_style)])
        # العمود الأيمن للتسمية، الأيسر للقيمة
        t = Table(rows, colWidths=[content_width * 0.62, content_width * 0.38])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D9E2EC')),
            ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#F0F4F8')),
        ]))
        return t

    for idx, app in enumerate(applications, 1):
        ans = answers_map.get(app.id, {})
        exps = experiences_map.get(app.id, [])
        decision = hiring_map.get(app.id)

        block = []

        # عنوان الطلب (شريط ملون)
        title_text = f"طلب رقم {app.id}  -  {app.applied_position or ''}  ({STATUS_LABELS.get(app.status, app.status)})"
        title_tbl = Table([[P(title_text, app_title_style)]], colWidths=[content_width])
        title_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1F4E79')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        block.append(title_tbl)
        block.append(Spacer(1, 0.15 * cm))

        # المعلومات الأساسية
        basic_pairs = [
            ('الوظيفة المتقدم لها', app.applied_position or ''),
            ('الحالة', STATUS_LABELS.get(app.status, app.status)),
            ('الفرع', app.branch.name if app.branch else ''),
            ('القسم', app.department.name if app.department else ''),
            ('التقييم العام', str(app.evaluation_score) if app.evaluation_score is not None else ''),
            ('تاريخ التقديم', app.created_at.strftime('%Y-%m-%d %H:%M') if app.created_at else ''),
            ('أُدخل بواسطة', app.creator.username if app.creator else ''),
        ]
        if app.status == 'rejected' and app.rejection_reason:
            basic_pairs.append(('سبب الرفض', app.rejection_reason))
        block.append(P("المعلومات الأساسية", section_title_style))
        block.append(Spacer(1, 0.1 * cm))
        block.append(kv_table(basic_pairs))
        block.append(Spacer(1, 0.25 * cm))

        # الإجابات الديناميكية مجمّعة حسب الأقسام
        for section, fields in sections_with_fields:
            display_fields = [f for f in fields if f.field_type != 'experience_group']
            section_pairs = []
            for field in display_fields:
                val = ans.get(field.id, '')
                if val is None or str(val).strip() == '':
                    continue
                section_pairs.append((field.label, _fmt(val)))
            if not section_pairs:
                continue
            block.append(P(section.name, section_title_style))
            block.append(Spacer(1, 0.1 * cm))
            block.append(kv_table(section_pairs))
            block.append(Spacer(1, 0.25 * cm))

        # جدول الخبرات السابقة (الأعمال السابقة)
        if exps:
            block.append(P("الخبرات / الأعمال السابقة", section_title_style))
            block.append(Spacer(1, 0.1 * cm))

            exp_header = [P('#', th_style)] + [P(lbl, th_style) for _key, lbl in EXP_SUB_HEADERS]
            exp_rows = [exp_header]
            for ei, exp in enumerate(exps, 1):
                row = [P(str(ei), td_style)]
                for key, _lbl in EXP_SUB_HEADERS:
                    row.append(P(_fmt(getattr(exp, key)), td_style))
                exp_rows.append(row)

            # عرض الأعمدة: رقم صغير + توزيع الباقي
            n_sub = len(EXP_SUB_HEADERS)
            num_w = 0.7 * cm
            rest = content_width - num_w
            col_widths = [num_w] + [rest / n_sub] * n_sub

            exp_tbl = Table(exp_rows, colWidths=col_widths, repeatRows=1)
            exp_tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
                ('FONTNAME', (0, 0), (-1, -1), arabic_font),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EEF6EE')]),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BBBBBB')),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            block.append(exp_tbl)
            block.append(Spacer(1, 0.25 * cm))

        # قرار التعيين (للمقبولين)
        if decision:
            hiring_pairs = [(lbl, _fmt(val)) for lbl, val in _hiring_columns(decision)
                            if _fmt(val).strip() != '']
            if hiring_pairs:
                block.append(P("قرار التعيين", section_title_style))
                block.append(Spacer(1, 0.1 * cm))
                ht = kv_table(hiring_pairs)
                ht.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), arabic_font),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ('LEFTPADDING', (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                    ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E8D9B5')),
                    ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#FBF6EA')),
                ]))
                block.append(ht)
                block.append(Spacer(1, 0.25 * cm))

        block.append(HRFlowable(width="100%", thickness=1.2,
                                color=colors.HexColor('#1F4E79'), spaceBefore=2, spaceAfter=10))

        # نحافظ على عنوان الطلب + معلوماته الأساسية معاً، والباقي يتدفق طبيعياً
        story.append(KeepTogether(block[:4]))
        story.extend(block[4:])

    doc.build(story)
    output.seek(0)
    return output
